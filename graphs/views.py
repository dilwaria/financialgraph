from django.shortcuts import render,redirect

from django.core.urlresolvers import reverse

# from .models import User
# Create your views here.
from django.http import JsonResponse,HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from graphs.models import book1, book2, book3, book4, book5 ,book6
from django.db.models import Max,Min,Sum
import numpy 
from random import randint
import json 
from django.core import serializers

# importing forms 
from graphs.forms import PostInputFormBook1,PostInputFormBook2,PostInputFormBook3,PostInputFormBook4

def introduction(request,variable_id=1):
	# openpyxl.load_workbook()							
	workbook = load_workbook('Book1.xlsx')
	ws1 = workbook.get_sheet_by_name("Sheet7")
	# print workBookBySheet.cell(row=1, column=2).value 
	# print workBookBySheet.cell(row=1, column=3).value
	# print ws1.cell(row=2,column=2).value
	# for i in range(2,5002):
	# 	x = book6(value1=ws1.cell(row=i, column=2).value,value2=ws1.cell(row=i, column=3).value,value3=ws1.cell(row=i, column=4).value,value4=ws1.cell(row=i, column=5).value,value5=ws1.cell(row=i, column=6).value,value6=ws1.cell(row=i, column=7).value,value7=ws1.cell(row=i, column=8).value,value8=ws1.cell(row=i, column=9).value,value9=ws1.cell(row=i, column=10).value,value10=ws1.cell(row=i, column=11).value  )
	# 	x.save()
		
	print variable_id
	if variable_id=="1": 
		# print "here"
		book = book1
		FORM = PostInputFormBook1()
	elif variable_id=="2":
		book =book2
		FORM = PostInputFormBook2()
	elif variable_id=="3":
		book=book3
		FORM = PostInputFormBook3()
	elif variable_id=="4":
		book = book4
		FORM = PostInputFormBook4()
	else:
		book =book1
		FORM = PostInputFormBook1()

	firstObs = book.objects.filter(table_type=variable_id).first()
	# for i in firstObs:
	# print firstObs

	lastObs = book.objects.filter(table_type=variable_id).last()
	# print lastObs
	MaxObs = book.objects.filter(table_type=variable_id).aggregate(Max('value1'),Max('value2'),Max('value3'),Max('value4'),Max('value5'),Max('value6'),Max('value7'),Max('value8'),Max('value9'),Max('value10'))
	MinObs = book.objects.filter(table_type=variable_id).aggregate(Min('value1'),Min('value2'),Min('value3'),Min('value4'),Min('value5'),Min('value6'),Min('value7'),Min('value8'),Min('value9'),Min('value10'))
	SumObs = book.objects.filter(table_type=variable_id).aggregate(Sum('value1'),Sum('value2'),Sum('value3'),Sum('value4'),Sum('value5'),Sum('value6'),Sum('value7'),Sum('value8'),Sum('value9'),Sum('value10'))
	
	largests = []
	smallests = []
	for i in range(1,11):
		target = "-value"+str(i)
		targetdesc = "value"+str(i)
		nlargest = book.objects.filter(table_type=variable_id).order_by(target)[0]
		nsmallest = book.objects.filter(table_type=variable_id).order_by(targetdesc)[0]
		largests.append(nlargest)
		smallests.append(nsmallest)


	# print smallests
	# for i in smallests:
	# 	print i.value1,i.value2,i.value3,i.value4,i.value5,i.value6,i.value7,i.value8,i.value9,i.value10


	# for i in range(0,10):
	# 	# targetdesc = "value"+str(i)
	# 	print smallests[i].value1

	# print smallests[0].value1
	# nsmallest = 

	finalarray =[]
	for i in range(1,11):
		# print i
		string = "value"+str(i)
		f = book.objects.values(string)
		emptyarray = []
		for i in f:
			emptyarray.append(i[string])

		xc = numpy.array(emptyarray)
		yc = numpy.transpose(emptyarray)
		finalarray.append(numpy.dot(xc,yc))

	# print finalarray

	if request.method == "POST":

		if variable_id=="2": 
			print "here"
			form = PostInputFormBook2(request.POST)
		elif variable_id=="3":
			form = PostInputFormBook3(request.POST)
		elif variable_id=="4":
			form = PostInputFormBook4(request.POST)
		else:
			form = PostInputFormBook1(request.POST)

		if form.is_valid():
			post = form.save()
			
			post.save()
			return redirect('introduction',variable_id)	

	GET_CONTEXT = {"firstobs": firstObs,
					"lastobs": lastObs,
					"maximum": MaxObs,
					"minimum": MinObs,
					"sum":SumObs,
					"matrix": finalarray,
					"form": FORM,
					"loop": range(1,100),
					"variable":variable_id,
					"largests": largests,
					"smallests": smallests,}
	return render(request, 'graphs/introductionagain.html', GET_CONTEXT)



def insertDataFunction():
	for i in range(2,5002):
		x = book4(value1=ws1.cell(row=i, column=2).value,value2=ws1.cell(row=i, column=3).value,value3=ws1.cell(row=i, column=4).value,value4=ws1.cell(row=i, column=5).value,value5=ws1.cell(row=i, column=6).value,value6=ws1.cell(row=i, column=7).value,value7=ws1.cell(row=i, column=8).value,value8=ws1.cell(row=i, column=9).value,value9=ws1.cell(row=i, column=10).value,value10=ws1.cell(row=i, column=11).value  )
		x.save()
 

# order = 1  ! largest
# order = 2   ! smallest 
def ajaxfire(request,variable_id,order,number):
	# print variable_id

	if request.method == "GET" and request.is_ajax():
		variable_id = request.GET['variable']
		number =  request.GET['number']

	# print number, variable_id
	if variable_id=="1": 
		book = book1
	elif variable_id=="2":
		book =book2	
	elif variable_id=="3":
		book=book3
	elif variable_id=="4":
		book = book4
	else:
		book =book1

	largests = []
	if order ==1 :
		target = "value"+str(i)
	elif order ==2 :
		target = "-value"+str(i)
	# smallests = []
	numb = int(number)
	for i in range(1,11):
		target = "value"+str(i)
		# print target
		nlargest = book.objects.all().order_by(target)[numb]
		largests.append(nlargest)
		# smallests.append(nsmallest)
	# print largests
	dataserialized  = serializers.serialize("json", largests)
	
	
	return  HttpResponse(dataserialized)




def graphing(request,variable_id=1):


	# print variable_id
	if variable_id=="1": 
		# print "here"
		book = book1
		
	elif variable_id=="2":
		book =book2
		
	elif variable_id=="3":
		book=book3
		
	elif variable_id=="4":
		book = book4
		
	else:
		book =book1
		

	firstObs = book.objects.first()
	lastObs = book.objects.last()
	MaxObs = book.objects.all().aggregate(Max('value1'),Max('value2'),Max('value3'),Max('value4'),Max('value5'),Max('value6'),Max('value7'),Max('value8'),Max('value9'),Max('value10'))
	MinObs = book.objects.all().aggregate(Min('value1'),Min('value2'),Min('value3'),Min('value4'),Min('value5'),Min('value6'),Min('value7'),Min('value8'),Min('value9'),Min('value10'))
	SumObs = book.objects.all().aggregate(Sum('value1'),Sum('value2'),Sum('value3'),Sum('value4'),Sum('value5'),Sum('value6'),Sum('value7'),Sum('value8'),Sum('value9'),Sum('value10'))
	
	finalarray =[]
	for i in range(1,11):
		# print i
		string = "value"+str(i)
		f = book.objects.values(string)
		emptyarray = []
		for i in f:
			emptyarray.append(i[string])

		xc = numpy.array(emptyarray)
		yc = numpy.transpose(emptyarray)
		finalarray.append(numpy.dot(xc,yc))

	# print finalarray


	GET_CONTEXT = {"firstobs": firstObs,
					"lastobs": lastObs,
					"maximum": MaxObs,
					"minimum": MinObs,
					"sum":SumObs,
					"matrix": finalarray,
					# "form": FORM,
					"loop": range(100),
					"variable":variable_id}


	# GET_CONTEXT = {}
	return render(request, 'graphs/graphing.html', GET_CONTEXT)


